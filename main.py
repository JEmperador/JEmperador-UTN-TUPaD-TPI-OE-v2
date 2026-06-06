"""
Bot de Telegram - Soporte Técnico Nivel 1
TeleNet S.A. - Proveedor de Internet y Telefonía
"""

import logging
import re
import os
import time
from datetime import datetime
from dotenv import load_dotenv

import openpyxl
import asyncio

from telegram import Update
from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    filters,
    ContextTypes,
)

# ==========================================================
# CONFIGURACIÓN
# ==========================================================

load_dotenv()
TOKEN = os.getenv("TOKEN")

EXCEL_PATH = "Datos/soporte.xlsx"
EMAIL_SOPORTE = "soporte@telenet.com.ar"
TIMEOUT_SEGUNDOS = 180
MAX_INTENTOS_NUMERO = 3

# Categorías de problemas (código: (nombre, soluciones))
CATEGORIAS = {
    "1": (
        "Conectividad (sin internet)",
        "• Reiniciá el router desenchufándolo 30 segundos.\n"
        "• Verificá que todos los cables estén bien conectados.\n"
        "• Consultá si hay un corte en tu zona en nuestra app.",
    ),
    "2": (
        "Velocidad lenta",
        "• Reiniciá el router y esperá 2 minutos.\n"
        "• Verificá cuántos dispositivos están conectados.\n"
        "• Realizá un test de velocidad en fast.com.",
    ),
    "3": (
        "Telefonía (sin línea)",
        "• Verificá que el cable del teléfono esté bien conectado al módem.\n"
        "• Reiniciá el módem y esperá que reconecte.\n"
        "• Probá con otro teléfono si tenés.",
    ),
    "4": (
        "Facturación y pagos",
        "• Revisá tu bandeja de entrada por el email de factura.\n"
        "• Ingresá a nuestra app para ver el estado de tu cuenta.\n"
        "• Verificá que tu medio de pago esté vigente.",
    ),
    "5": (
        "Otro problema",
        None,  # Deriva directo a ticket sin mostrar soluciones
    ),
}

# ==========================================================
# ESTADOS (Máquina de Estados)
# ==========================================================

ESTADO_ESPERANDO_NUMERO = "ESPERANDO_NUMERO"
ESTADO_ESPERANDO_CATEGORIA = "ESPERANDO_CATEGORIA"
ESTADO_ESPERANDO_CONFIRMACION = "ESPERANDO_CONFIRMACION"
ESTADO_ESPERANDO_EMAIL = "ESPERANDO_EMAIL"
ESTADO_CERRADO = "CERRADO"
ESTADO_TICKET_GENERADO = "TICKET_GENERADO"
ESTADO_FALLIDO = "FALLIDO"

# ==========================================================
# LOGGING
# ==========================================================

logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO,
)
logger = logging.getLogger(__name__)

# ==========================================================
# INICIALIZACIÓN DEL EXCEL
# ==========================================================


def inicializar_excel():
    """
    Crea soporte.xlsx con las tres hojas definidas en el diccionario de datos
    si el archivo no existe. Si existe, no lo toca.

    Hojas:
      - Clientes  → base de clientes habilitados (precargada con datos de ejemplo)
      - Sesiones  → estado actual de cada conversación activa
      - Historial → log completo de interacciones
    """
    if os.path.exists(EXCEL_PATH):
        return

    wb = openpyxl.Workbook()

    # ── Hoja Clientes ──────────────────────────────────────
    ws_clientes = wb.active
    ws_clientes.title = "Clientes"
    ws_clientes.append(
        [
            "numero_cliente",
            "nombre_completo",
            "telefono",
            "email",
            "servicio_contratado",
            "ticket_asociado",
            "fecha_alta",
        ]
    )
    # Datos de ejemplo para poder probar el bot
    ws_clientes.append(
        [
            "CLI-0001",
            "García, Laura",
            "+54 11 4567-8901",
            "laura@gmail.com",
            "Internet Fibra 100Mbps",
            "",
            "15-03-2023",
        ]
    )
    ws_clientes.append(
        [
            "CLI-0002",
            "Martínez, Carlos",
            "+54 11 5678-9012",
            "carlos@gmail.com",
            "Internet + Telefonía",
            "",
            "20-07-2022",
        ]
    )
    ws_clientes.append(
        [
            "CLI-0003",
            "López, Ana",
            "+54 11 6789-0123",
            "ana@hotmail.com",
            "Internet Fibra 50Mbps",
            "",
            "01-01-2024",
        ]
    )

    # ── Hoja Sesiones ──────────────────────────────────────
    ws_sesiones = wb.create_sheet("Sesiones")
    ws_sesiones.append(
        [
            "telegram_id",
            "numero_cliente",
            "estado_actual",
            "intentos_numero",
            "categoria_seleccionada",
            "email_contacto",
            "ticket_id",
            "ultimo_contacto",
        ]
    )

    # ── Hoja Historial ─────────────────────────────────────
    ws_historial = wb.create_sheet("Historial")
    ws_historial.append(
        [
            "id_interaccion",
            "telegram_id",
            "numero_cliente",
            "fecha_hora",
            "accion",
            "detalle",
            "resultado",
            "ticket_id",
        ]
    )

    wb.save(EXCEL_PATH)
    logger.info(f"Excel creado en: {EXCEL_PATH}")


# ==========================================================
# CAPA DE ACCESO A DATOS
# ==========================================================


def _timestamp() -> str:
    return datetime.now().strftime("%d-%m-%Y %H:%M")


def cliente_existe(numero_cliente: str) -> bool:
    """Consulta la hoja Clientes. Devuelve True si el número existe. (Gateway 1)"""
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb["Clientes"]
        for fila in ws.iter_rows(min_row=2, values_only=True):
            if str(fila[0]).strip() == numero_cliente.strip():
                return True
    except Exception as e:
        logger.error(f"Error al verificar cliente: {e}")
    return False


def obtener_proximo_ticket_id() -> str:
    """
    Genera el próximo ticket_id correlativo en formato TKT-XXXX
    leyendo el máximo existente en la hoja Historial.
    """
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb["Historial"]
        maximo = 0
        for fila in ws.iter_rows(min_row=2, values_only=True):
            ticket = fila[7]  # columna ticket_id
            if ticket and str(ticket).startswith("TKT-"):
                numero = int(str(ticket).replace("TKT-", ""))
                if numero > maximo:
                    maximo = numero
        return f"TKT-{(maximo + 1):04d}"
    except Exception as e:
        logger.error(f"Error al generar ticket_id: {e}")
        return "TKT-0001"


def upsert_sesion(telegram_id: int, datos: dict):
    """
    Crea o actualiza la fila de sesión del usuario en la hoja Sesiones.
    Si el telegram_id ya existe, actualiza. Si no, inserta.
    """
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb["Sesiones"]
        for fila in ws.iter_rows(min_row=2):
            if fila[0].value == telegram_id:
                fila[0].value = telegram_id
                fila[1].value = datos.get("numero_cliente", fila[1].value)
                fila[2].value = datos.get("estado_actual", fila[2].value)
                fila[3].value = datos.get("intentos_numero", fila[3].value)
                fila[4].value = datos.get("categoria_seleccionada", fila[4].value)
                fila[5].value = datos.get("email_contacto", fila[5].value)
                fila[6].value = datos.get("ticket_id", fila[6].value)
                fila[7].value = _timestamp()
                wb.save(EXCEL_PATH)
                return
        # No existe → insertar nueva fila
        ws.append(
            [
                telegram_id,
                datos.get("numero_cliente", ""),
                datos.get("estado_actual", ""),
                datos.get("intentos_numero", 0),
                datos.get("categoria_seleccionada", ""),
                datos.get("email_contacto", ""),
                datos.get("ticket_id", ""),
                _timestamp(),
            ]
        )
        wb.save(EXCEL_PATH)
    except Exception as e:
        logger.error(f"Error en upsert_sesion: {e}")


def actualizar_ticket_en_cliente(numero_cliente: str, ticket_id: str):
    """Actualiza el campo ticket_asociado en la hoja Clientes. (RN-06)"""
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb["Clientes"]
        for fila in ws.iter_rows(min_row=2):
            if str(fila[0].value).strip() == numero_cliente.strip():
                fila[5].value = ticket_id  # columna ticket_asociado
                wb.save(EXCEL_PATH)
                return
    except Exception as e:
        logger.error(f"Error al actualizar ticket en cliente: {e}")


def registrar_historial(
    telegram_id: int,
    accion: str,
    resultado: str,
    numero_cliente: str = "",
    detalle: str = "",
    ticket_id: str = "",
):
    """
    Inserta una fila en la hoja Historial. (RN-07)
    Toda acción queda registrada, sin excepción.
    """
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb["Historial"]
        # Calcular próximo id_interaccion
        max_id = 0
        for fila in ws.iter_rows(min_row=2, values_only=True):
            if fila[0] and isinstance(fila[0], int):
                max_id = max(max_id, fila[0])
        ws.append(
            [
                max_id + 1,
                telegram_id,
                numero_cliente,
                _timestamp(),
                accion,
                detalle,
                resultado,
                ticket_id,
            ]
        )
        wb.save(EXCEL_PATH)
    except Exception as e:
        logger.error(f"Error al registrar historial: {e}")


# ==========================================================
# VALIDACIONES
# ==========================================================


def validar_numero_cliente(texto: str) -> bool:
    """Formato válido: CLI-XXXX (ej. CLI-0001)"""
    return bool(re.match(r"^CLI-\d{4}$", texto.strip().upper()))


def validar_email(texto: str) -> bool:
    return bool(re.match(r"^[\w\.-]+@[\w\.-]+\.\w{2,}$", texto.strip()))


def validar_categoria(texto: str) -> bool:
    return texto.strip() in CATEGORIAS


# ==========================================================
# TIMEOUT
# ==========================================================


def actualizar_timestamp(context: ContextTypes.DEFAULT_TYPE):
    context.user_data["ultimo_mensaje"] = time.time()


async def verificar_timeout(update: Update, context: ContextTypes.DEFAULT_TYPE) -> bool:
    ultimo = context.user_data.get("ultimo_mensaje")
    if ultimo and (time.time() - ultimo) > TIMEOUT_SEGUNDOS:
        telegram_id = update.effective_user.id
        registrar_historial(
            telegram_id=telegram_id,
            accion="TIMEOUT",
            resultado="FALLIDO",
            numero_cliente=context.user_data.get("numero_cliente", ""),
        )
        context.user_data.clear()
        await update.message.reply_text(
            "⏱ Tu sesión expiró por inactividad (3 minutos).\n"
            "Escribí /start para iniciar una nueva consulta.\n\n"
            f"📧 Consultas: {EMAIL_SOPORTE}"
        )
        return True
    actualizar_timestamp(context)
    return False


async def verificar_inactividad(context: ContextTypes.DEFAULT_TYPE):
    """Job periódico que detecta sesiones inactivas y las cierra."""
    chat_id = context.job.chat_id
    user_data = context.application.user_data.get(chat_id, {})
    ultimo = user_data.get("ultimo_mensaje")
    estado = user_data.get("estado")

    if not ultimo or estado in [
        None,
        ESTADO_CERRADO,
        ESTADO_TICKET_GENERADO,
        ESTADO_FALLIDO,
    ]:
        return

    if (time.time() - ultimo) > TIMEOUT_SEGUNDOS:
        registrar_historial(
            telegram_id=chat_id,
            accion="TIMEOUT",
            resultado="FALLIDO",
            numero_cliente=user_data.get("numero_cliente", ""),
        )
        user_data.clear()
        await context.bot.send_message(
            chat_id=chat_id,
            text=(
                "⏱ Tu sesión expiró por inactividad (3 minutos).\n"
                "Escribí /start para iniciar una nueva consulta.\n\n"
                f"📧 Consultas: {EMAIL_SOPORTE}"
            ),
        )


# ==========================================================
# HELPERS
# ==========================================================


def cancelar_jobs(context: ContextTypes.DEFAULT_TYPE, chat_id: int):
    for job in context.job_queue.get_jobs_by_name(str(chat_id)):
        job.schedule_removal()


def texto_categorias() -> str:
    lines = ["Seleccioná el número de tu problema:\n"]
    for codigo, (nombre, _) in CATEGORIAS.items():
        lines.append(f"  *{codigo}* — {nombre}")
    return "\n".join(lines)


# ==========================================================
# HANDLERS — Comandos
# ==========================================================


async def cmd_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    cancelar_jobs(context, update.effective_chat.id)
    context.user_data.clear()
    context.user_data["estado"] = ESTADO_ESPERANDO_NUMERO
    context.user_data["intentos_numero"] = 0
    actualizar_timestamp(context)

    telegram_id = update.effective_user.id

    upsert_sesion(
        telegram_id,
        {
            "estado_actual": ESTADO_ESPERANDO_NUMERO,
            "intentos_numero": 0,
        },
    )
    registrar_historial(
        telegram_id=telegram_id,
        accion="INICIO_SESION",
        resultado="EXITOSO",
    )

    context.job_queue.run_repeating(
        verificar_inactividad,
        interval=30,
        first=30,
        chat_id=update.effective_chat.id,
        name=str(update.effective_chat.id),
    )

    await update.message.reply_text(
        "👋 ¡Bienvenido al Soporte Técnico de *TeleNet S.A.*!\n\n"
        "Voy a ayudarte a resolver tu problema paso a paso.\n\n"
        "Para comenzar, ingresá tu *número de cliente*.\n"
        "Formato: *CLI-XXXX* (lo encontrás en tu factura)\n\n"
        "Podés escribir /cancelar en cualquier momento para salir.",
        parse_mode="Markdown",
    )


async def cmd_cancelar(update: Update, context: ContextTypes.DEFAULT_TYPE):
    cancelar_jobs(context, update.effective_chat.id)
    telegram_id = update.effective_user.id
    registrar_historial(
        telegram_id=telegram_id,
        accion="CANCELADO",
        resultado="FALLIDO",
        numero_cliente=context.user_data.get("numero_cliente", ""),
    )
    upsert_sesion(
        telegram_id,
        {
            "estado_actual": ESTADO_FALLIDO,
            "numero_cliente": context.user_data.get("numero_cliente", ""),
        },
    )
    context.user_data.clear()
    await update.message.reply_text(
        "❌ Consulta cancelada.\n"
        "Escribí /start cuando quieras iniciar una nueva.\n\n"
        f"📧 Consultas: {EMAIL_SOPORTE}"
    )


async def cmd_ayuda(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "📋 *Comandos disponibles:*\n\n"
        "/start — Iniciar una nueva consulta de soporte\n"
        "/cancelar — Cancelar la consulta actual\n"
        "/estado — Ver en qué paso estás\n"
        "/ayuda — Mostrar esta guía\n\n"
        f"📧 Consultas: {EMAIL_SOPORTE}",
        parse_mode="Markdown",
    )


async def cmd_estado(update: Update, context: ContextTypes.DEFAULT_TYPE):
    estado = context.user_data.get("estado")
    mensajes = {
        ESTADO_ESPERANDO_NUMERO: "Estoy esperando tu número de cliente (CLI-XXXX).",
        ESTADO_ESPERANDO_CATEGORIA: "Estoy esperando que selecciones la categoría del problema (1 al 5).",
        ESTADO_ESPERANDO_CONFIRMACION: "Estoy esperando que confirmes si el problema se resolvió (sí/no).",
        ESTADO_ESPERANDO_EMAIL: "Estoy esperando tu email de contacto para el técnico.",
        ESTADO_CERRADO: "Tu consulta fue cerrada exitosamente.",
        ESTADO_TICKET_GENERADO: "Hay un ticket abierto. El técnico se contactará pronto.",
        ESTADO_FALLIDO: "La sesión terminó. Escribí /start para iniciar de nuevo.",
    }
    texto = mensajes.get(
        estado, "No hay una consulta activa. Escribí /start para comenzar."
    )
    await update.message.reply_text(f"📍 {texto}")


# ==========================================================
# HANDLER — Mensajes de texto
# ==========================================================


async def manejar_mensaje(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if await verificar_timeout(update, context):
        return

    estado = context.user_data.get("estado")
    texto = update.message.text.strip()
    telegram_id = update.effective_user.id

    # Sin sesión activa
    if not estado:
        await update.message.reply_text(
            "No hay una consulta activa.\n" "Escribí /start para comenzar."
        )
        return

    # Estados finales: no procesar más input
    if estado in [ESTADO_CERRADO, ESTADO_TICKET_GENERADO]:
        await update.message.reply_text(
            "Tu consulta ya fue procesada.\n" "Escribí /start para iniciar una nueva.",
        )
        return

    if estado == ESTADO_FALLIDO:
        await update.message.reply_text(
            "La sesión está cerrada por intentos fallidos.\n"
            "Escribí /start para intentar nuevamente."
        )
        return

    # ── ESPERANDO_NUMERO ────────────────────────────────────
    if estado == ESTADO_ESPERANDO_NUMERO:
        numero = texto.upper()

        if not validar_numero_cliente(numero):
            # Formato inválido — cuenta como intento
            context.user_data["intentos_numero"] += 1
            intentos = context.user_data["intentos_numero"]
            registrar_historial(
                telegram_id=telegram_id,
                accion="NUMERO_INVALIDO",
                resultado="FALLIDO",
                detalle=f"Formato inválido: {texto}. Intento {intentos}",
            )
            if intentos >= MAX_INTENTOS_NUMERO:
                context.user_data["estado"] = ESTADO_FALLIDO
                upsert_sesion(
                    telegram_id,
                    {"estado_actual": ESTADO_FALLIDO, "intentos_numero": intentos},
                )
                await update.message.reply_text(
                    "❌ Superaste el límite de intentos.\n"
                    "Escribí /start cuando estés listo para intentarlo de nuevo.\n\n"
                    f"📧 Consultas: {EMAIL_SOPORTE}"
                )
                return
            restantes = MAX_INTENTOS_NUMERO - intentos
            await update.message.reply_text(
                f"❌ Formato incorrecto. El número debe tener el formato *CLI-XXXX*.\n"
                f"Ejemplo: *CLI-0042*\n\n"
                f"⚠️ Intentos restantes: *{restantes}*",
                parse_mode="Markdown",
            )
            return

        # Formato correcto → verificar en BD (Gateway 1)
        if not cliente_existe(numero):
            context.user_data["intentos_numero"] += 1
            intentos = context.user_data["intentos_numero"]
            registrar_historial(
                telegram_id=telegram_id,
                accion="NUMERO_INVALIDO",
                resultado="FALLIDO",
                detalle=f"No encontrado en BD: {numero}. Intento {intentos}",
            )
            if intentos >= MAX_INTENTOS_NUMERO:
                context.user_data["estado"] = ESTADO_FALLIDO
                upsert_sesion(
                    telegram_id,
                    {"estado_actual": ESTADO_FALLIDO, "intentos_numero": intentos},
                )
                await update.message.reply_text(
                    "❌ Superaste el límite de intentos.\n"
                    "Escribí /start cuando estés listo para intentarlo de nuevo.\n\n"
                    f"📧 Consultas: {EMAIL_SOPORTE}"
                )
                return
            restantes = MAX_INTENTOS_NUMERO - intentos
            await update.message.reply_text(
                f"❌ El número *{numero}* no está registrado en nuestro sistema.\n"
                f"Verificá el número en tu factura e intentá de nuevo.\n\n"
                f"⚠️ Intentos restantes: *{restantes}*",
                parse_mode="Markdown",
            )
            return

        # Cliente válido → avanzar
        context.user_data["numero_cliente"] = numero
        context.user_data["estado"] = ESTADO_ESPERANDO_CATEGORIA
        upsert_sesion(
            telegram_id,
            {
                "numero_cliente": numero,
                "estado_actual": ESTADO_ESPERANDO_CATEGORIA,
                "intentos_numero": context.user_data["intentos_numero"],
            },
        )
        registrar_historial(
            telegram_id=telegram_id,
            accion="NUMERO_VALIDADO",
            resultado="EXITOSO",
            numero_cliente=numero,
        )
        await update.message.reply_text(
            f"✅ Cliente *{numero}* verificado.\n\n" f"{texto_categorias()}",
            parse_mode="Markdown",
        )
        return

    # ── ESPERANDO_CATEGORIA ─────────────────────────────────
    if estado == ESTADO_ESPERANDO_CATEGORIA:
        numero_cliente = context.user_data.get("numero_cliente", "")

        if not validar_categoria(texto):
            await update.message.reply_text(
                "❌ Opción inválida. Por favor ingresá un número del *1 al 5*.\n\n"
                f"{texto_categorias()}",
                parse_mode="Markdown",
            )
            return

        nombre_categoria, soluciones = CATEGORIAS[texto]
        context.user_data["categoria_seleccionada"] = nombre_categoria
        registrar_historial(
            telegram_id=telegram_id,
            accion="CATEGORIA_SELECCIONADA",
            resultado="EXITOSO",
            numero_cliente=numero_cliente,
            detalle=f"Categoría: {nombre_categoria}",
        )

        # Categoría 5 "Otro problema" deriva directo a ticket sin mostrar soluciones
        if soluciones is None:
            context.user_data["estado"] = ESTADO_ESPERANDO_EMAIL
            upsert_sesion(
                telegram_id,
                {
                    "estado_actual": ESTADO_ESPERANDO_EMAIL,
                    "categoria_seleccionada": nombre_categoria,
                },
            )
            await update.message.reply_text(
                "Entendido. Para derivar tu caso a un técnico necesito tu *email de contacto*.\n"
                "Ejemplo: *tucorreo@gmail.com*",
                parse_mode="Markdown",
            )
            return

        # Mostrar soluciones frecuentes (Gateway 2 próximo paso)
        context.user_data["estado"] = ESTADO_ESPERANDO_CONFIRMACION
        upsert_sesion(
            telegram_id,
            {
                "estado_actual": ESTADO_ESPERANDO_CONFIRMACION,
                "categoria_seleccionada": nombre_categoria,
            },
        )
        registrar_historial(
            telegram_id=telegram_id,
            accion="SOLUCION_MOSTRADA",
            resultado="EXITOSO",
            numero_cliente=numero_cliente,
            detalle=f"Categoría: {nombre_categoria}",
        )
        await update.message.reply_text(
            f"📋 *{nombre_categoria}* — Pasos a seguir:\n\n"
            f"{soluciones}\n\n"
            "¿El problema se resolvió? Respondé *sí* o *no*.",
            parse_mode="Markdown",
        )
        return

    # ── ESPERANDO_CONFIRMACION (Gateway 2) ──────────────────
    if estado == ESTADO_ESPERANDO_CONFIRMACION:
        numero_cliente = context.user_data.get("numero_cliente", "")
        respuesta = texto.lower().strip()

        if respuesta in ["si", "sí", "s"]:
            context.user_data["estado"] = ESTADO_CERRADO
            upsert_sesion(telegram_id, {"estado_actual": ESTADO_CERRADO})
            registrar_historial(
                telegram_id=telegram_id,
                accion="PROBLEMA_RESUELTO",
                resultado="EXITOSO",
                numero_cliente=numero_cliente,
            )
            cancelar_jobs(context, update.effective_chat.id)
            await update.message.reply_text(
                "✅ ¡Nos alegra que se haya resuelto!\n\n"
                "Tu caso fue cerrado exitosamente.\n"
                "Escribí /start si necesitás hacer otra consulta.\n\n"
                f"📧 Consultas: {EMAIL_SOPORTE}"
            )
            return

        if respuesta in ["no", "n"]:
            context.user_data["estado"] = ESTADO_ESPERANDO_EMAIL
            upsert_sesion(telegram_id, {"estado_actual": ESTADO_ESPERANDO_EMAIL})
            await update.message.reply_text(
                "Entendido. Para derivar tu caso a un técnico necesito tu *email de contacto*.\n"
                "Ejemplo: *tucorreo@gmail.com*",
                parse_mode="Markdown",
            )
            return

        # Respuesta inválida
        await update.message.reply_text(
            "❌ Por favor respondé únicamente *sí* o *no*.",
            parse_mode="Markdown",
        )
        return

    # ── ESPERANDO_EMAIL ─────────────────────────────────────
    if estado == ESTADO_ESPERANDO_EMAIL:
        numero_cliente = context.user_data.get("numero_cliente", "")

        if not validar_email(texto):
            await update.message.reply_text(
                "❌ El email ingresado no es válido.\n"
                "Por favor ingresá un email correcto.\n"
                "Ejemplo: *tucorreo@gmail.com*",
                parse_mode="Markdown",
            )
            return

        # Email válido → generar ticket
        ticket_id = obtener_proximo_ticket_id()
        context.user_data["email_contacto"] = texto
        context.user_data["ticket_id"] = ticket_id
        context.user_data["estado"] = ESTADO_TICKET_GENERADO

        upsert_sesion(
            telegram_id,
            {
                "estado_actual": ESTADO_TICKET_GENERADO,
                "email_contacto": texto,
                "ticket_id": ticket_id,
            },
        )
        actualizar_ticket_en_cliente(numero_cliente, ticket_id)
        registrar_historial(
            telegram_id=telegram_id,
            accion="EMAIL_REGISTRADO",
            resultado="EXITOSO",
            numero_cliente=numero_cliente,
            detalle=f"Email: {texto}",
        )
        registrar_historial(
            telegram_id=telegram_id,
            accion="TICKET_GENERADO",
            resultado="EXITOSO",
            numero_cliente=numero_cliente,
            detalle=f"Categoría: {context.user_data.get('categoria_seleccionada', '')}",
            ticket_id=ticket_id,
        )

        cancelar_jobs(context, update.effective_chat.id)
        await update.message.reply_text(
            f"🎫 *Ticket generado: {ticket_id}*\n\n"
            f"Un técnico especialista revisará tu caso y se contactará "
            f"al email *{texto}* a la brevedad.\n\n"
            f"Guardá tu número de ticket por si necesitás hacer seguimiento.\n\n"
            f"📧 Consultas: {EMAIL_SOPORTE}",
            parse_mode="Markdown",
        )
        return


# ==========================================================
# HANDLER — Mensajes no soportados
# ==========================================================


async def manejar_no_soportado(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if await verificar_timeout(update, context):
        return
    await update.message.reply_text(
        "❌ Este tipo de mensaje no es válido en el proceso de soporte.\n"
        "Por favor respondé con texto.\n\n"
        "Escribí /estado para ver en qué paso estás."
    )


# ==========================================================
# PUNTO DE ENTRADA
# ==========================================================


def main():
    asyncio.set_event_loop(asyncio.new_event_loop())
    
    inicializar_excel()

    app = Application.builder().token(TOKEN).build()

    app.add_handler(CommandHandler("start", cmd_start))
    app.add_handler(CommandHandler("cancelar", cmd_cancelar))
    app.add_handler(CommandHandler("ayuda", cmd_ayuda))
    app.add_handler(CommandHandler("estado", cmd_estado))

    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, manejar_mensaje))
    app.add_handler(
        MessageHandler(
            filters.PHOTO
            | filters.Document.ALL
            | filters.AUDIO
            | filters.VIDEO
            | filters.Sticker.ALL,
            manejar_no_soportado,
        )
    )

    logger.info("Bot iniciado. Esperando mensajes...")
    app.run_polling()


if __name__ == "__main__":
    main()
